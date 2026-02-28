#!/usr/bin/env python3
"""
Build an analytical spreadsheet for systematic quotation/allusion improvement.

Reads Hardy's biblical references, looks up verse texts, matches with phrase index,
and outputs to an Excel file with three tabs:
  1. "BoM Verses" - all BoM verses with Hardy refs and current phrase matches
  2. "Bible Verses" - all unique Bible verses with their texts
  3. "Allusion Language" - allusions that already have phrase matches (for review)
"""

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from bom_abbreviations import parse_bom_ref, ABBREV_TO_ID


# Mapping from Hardy's abbreviations to full book names as they appear in lds-scriptures.txt
HARDY_ABBREV_TO_NAME = {
    '1 Ne': '1 Nephi',
    '2 Ne': '2 Nephi',
    '3 Ne': '3 Nephi',
    '4 Ne': '4 Nephi',
    'Jac': 'Jacob',
    'Enos': 'Enos',
    'Jar': 'Jarom',
    'Omni': 'Omni',
    'WoM': 'Words of Mormon',
    'Mos': 'Mosiah',
    'Al': 'Alma',
    'Hel': 'Helaman',
    'Morm': 'Mormon',
    'Eth': 'Ether',
    'Ether': 'Ether',
    'Moro': 'Moroni',
}

# BoM book ordering for sorting
BOM_BOOK_ORDER = [
    '1nephi', '2nephi', '3nephi', '4nephi',
    'jacob', 'enos', 'jarom', 'omni', 'words-of-mormon',
    'mosiah', 'alma', 'helaman', 'mormon', 'ether', 'moroni'
]


def load_lds_scriptures():
    """Load lds-scriptures.txt into a dict {(book, chapter, verse): text}."""
    scriptures = {}
    with open('data/lds-scriptures.txt', 'r', encoding='utf-8') as f:
        for line in f:
            line = line.rstrip('\r\n')
            if not line:
                continue
            # Split on multiple spaces (or mixed whitespace)
            # Find the first occurrence of multiple spaces
            import re
            parts = re.split(r'\s{2,}', line, maxsplit=1)
            if len(parts) < 2:
                continue
            ref = parts[0]
            text = parts[1]

            # Parse reference: "Book Chapter:Verse"
            # Find last space to split book from chapter:verse
            last_space = ref.rfind(' ')
            if last_space == -1:
                continue
            book = ref[:last_space]
            verse_ref = ref[last_space+1:]
            if ':' not in verse_ref:
                continue
            chapter_str, verse_str = verse_ref.split(':', 1)
            try:
                chapter = int(chapter_str)
                verse = int(verse_str)
                scriptures[(book, chapter, verse)] = text
            except ValueError:
                continue
    return scriptures


def get_bom_verse_text(scriptures, book_id, chapter, verse):
    """Look up BoM verse text from scriptures dict. book_id like 'alma', 'mosiah', etc."""
    # Convert book_id back to full name
    id_to_name = {v: k for k, v in ABBREV_TO_ID.items()}
    # Handle special cases
    id_to_name['2nephi'] = '2 Nephi'
    id_to_name['3nephi'] = '3 Nephi'
    id_to_name['4nephi'] = '4 Nephi'
    id_to_name['words-of-mormon'] = 'Words of Mormon'

    # Map to full names
    book_mapping = {
        'alma': 'Alma',
        'mosiah': 'Mosiah',
        'helaman': 'Helaman',
        'mormon': 'Mormon',
        'ether': 'Ether',
        'moroni': 'Moroni',
        'jacob': 'Jacob',
        'enos': 'Enos',
        'jarom': 'Jarom',
        'omni': 'Omni',
        '1nephi': '1 Nephi',
        '2nephi': '2 Nephi',
        '3nephi': '3 Nephi',
        '4nephi': '4 Nephi',
        'words-of-mormon': 'Words of Mormon',
    }

    book_name = book_mapping.get(book_id)
    if not book_name:
        return None

    return scriptures.get((book_name, chapter, verse))


def get_bible_verse_text(scriptures, bible_ref):
    """Look up Bible verse text. bible_ref like 'Genesis 1.1' (dots to colons)."""
    # Convert dots to colons
    bible_ref_colon = bible_ref.replace('.', ':')
    # Parse
    last_space = bible_ref_colon.rfind(' ')
    if last_space == -1:
        return None
    book = bible_ref_colon[:last_space]
    verse_ref = bible_ref_colon[last_space+1:]
    try:
        chapter_str, verse_str = verse_ref.split(':', 1)
        chapter = int(chapter_str)
        verse = int(verse_str)
        return scriptures.get((book, chapter, verse))
    except (ValueError, AttributeError):
        return None


def parse_bible_ref(ref_str):
    """Parse a Bible reference like 'Isaiah 40.3' or 'Isaiah 52.1-3' or 'Isaiah 40.3, 5'."""
    # Split on comma for multiple verses
    parts = ref_str.split(',')
    verses = []
    for part in parts:
        part = part.strip()
        if '-' in part:
            # Range like "Isaiah 52.1-3"
            last_space = part.rfind(' ')
            book = part[:last_space]
            range_part = part[last_space+1:]
            if '.' not in range_part:
                continue
            chapter_str, verse_range = range_part.split('.', 1)
            try:
                chapter = int(chapter_str)
                range_parts = verse_range.split('-')
                start_verse = int(range_parts[0])
                end_verse = int(range_parts[1])
                for v in range(start_verse, end_verse + 1):
                    verses.append(f'{book} {chapter}:{v}')
            except (ValueError, IndexError):
                continue
        else:
            # Single verse
            last_space = part.rfind(' ')
            if last_space == -1:
                continue
            book = part[:last_space]
            verse_ref = part[last_space+1:]
            if '.' not in verse_ref:
                continue
            chapter_str, verse_str = verse_ref.split('.', 1)
            try:
                chapter = int(chapter_str)
                verse = int(verse_str)
                verses.append(f'{book} {chapter}:{v}' if (v := verse) else None)
            except ValueError:
                continue
    return verses


def get_bom_book_id(bom_ref):
    """Extract book_id from a parsed BoM ref tuple."""
    parsed = parse_bom_ref(bom_ref)
    if parsed:
        return parsed[0][0]
    return None


def get_bom_chapter_verse(bom_ref):
    """Extract (chapter, verse) from a parsed BoM ref. Returns first if multiple."""
    parsed = parse_bom_ref(bom_ref)
    if parsed:
        return (parsed[0][1], parsed[0][2])
    return (0, 0)


def load_hardy_data():
    """Load hardy_biblical_references.json."""
    with open('data/hardy_biblical_references.json', 'r') as f:
        return json.load(f)


def load_phrase_index():
    """Load hardy_phrase_index.json."""
    with open('data/hardy_phrase_index.json', 'r') as f:
        return json.load(f)


def get_phrase_matches_for_bom(phrase_index, bom_ref):
    """Get phrase matches for a BoM verse from the phrase index.

    Returns list of matched phrase texts, or empty list if none found.
    """
    parsed = parse_bom_ref(bom_ref)
    if not parsed:
        return []

    # phrase_index is nested by book_id, chapter, verse
    matches = []
    for book_id, chapter, verse in parsed:
        # Convert book_id to the key used in phrase_index
        # phrase_index uses keys like "alma", "mosiah", etc directly
        if book_id in phrase_index:
            chapter_data = phrase_index[book_id].get(str(chapter))
            if chapter_data:
                verse_data = chapter_data.get(str(verse))
                if verse_data and 'matches' in verse_data:
                    for match in verse_data['matches']:
                        matches.append(match['text'])

    return matches


def collect_bom_verses(hardy_data):
    """Collect all BoM verses (quotations + allusions, not cross-references).

    Returns list of (bom_ref, bible_ref, type) tuples.
    Only includes entries that can be parsed successfully.
    """
    verses = []
    for entry in hardy_data['entries']:
        if entry['type'] == 'cross-reference':
            continue
        # Only include if it can be parsed
        if parse_bom_ref(entry['bom_ref']):
            verses.append((entry['bom_ref'], entry['bible_ref'], entry['type']))
    return verses


def collect_unique_bible_refs(hardy_data):
    """Collect all unique Bible references."""
    refs = set()
    for entry in hardy_data['entries']:
        refs.add(entry['bible_ref'])
    return sorted(refs)


def collect_allusions_with_matches(hardy_data, phrase_index):
    """Collect allusions that have phrase matches.

    Returns list of (bible_ref, bom_ref, matched_phrases) tuples.
    Only includes entries that can be parsed successfully.
    """
    result = []
    seen = set()

    for entry in hardy_data['entries']:
        if entry['type'] != 'allusion':
            continue

        bom_ref = entry['bom_ref']
        bible_ref = entry['bible_ref']

        # Only process if it can be parsed
        if not parse_bom_ref(bom_ref):
            continue

        # Check if this verse has phrase matches
        matches = get_phrase_matches_for_bom(phrase_index, bom_ref)
        if matches:
            key = (bible_ref, bom_ref)
            if key not in seen:
                result.append((bible_ref, bom_ref, matches))
                seen.add(key)

    return result


def get_sorted_bom_verses(bom_verses):
    """Sort BoM verses by book order, then chapter, then verse."""
    def sort_key(item):
        bom_ref = item[0]
        book_id = get_bom_book_id(bom_ref)
        chapter, verse = get_bom_chapter_verse(bom_ref)

        if book_id not in BOM_BOOK_ORDER:
            book_idx = 999
        else:
            book_idx = BOM_BOOK_ORDER.index(book_id)

        return (book_idx, chapter, verse)

    return sorted(bom_verses, key=sort_key)


def auto_fit_column(ws, col_idx, max_width=80):
    """Auto-fit column width, capped at max_width."""
    col_letter = get_column_letter(col_idx)
    max_len = 0
    for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
        for cell in row:
            if cell.value:
                cell_len = len(str(cell.value))
                max_len = max(max_len, cell_len)

    adjusted_width = min(max_len + 2, max_width)
    ws.column_dimensions[col_letter].width = adjusted_width


def create_excel(output_path, bom_verses_data, bible_verses_data, allusions_with_matches_data):
    """Create the Excel workbook with three tabs."""
    wb = openpyxl.Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Define header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    # Tab 1: BoM Verses
    ws1 = wb.create_sheet('BoM Verses')
    headers1 = ['Type', 'Hardy Bible Ref', 'Hardy BoM Ref', 'BoM Verse Text', 'Current Phrase Match']
    for col_idx, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    for row_idx, (type_val, bible_ref, bom_ref, bom_text, phrases) in enumerate(bom_verses_data, 2):
        ws1.cell(row=row_idx, column=1, value=type_val)
        ws1.cell(row=row_idx, column=2, value=bible_ref)
        ws1.cell(row=row_idx, column=3, value=bom_ref)
        ws1.cell(row=row_idx, column=4, value=bom_text)
        ws1.cell(row=row_idx, column=5, value=phrases)

    # Auto-fit columns
    for col_idx in range(1, 6):
        auto_fit_column(ws1, col_idx)

    # Freeze top row
    ws1.freeze_panes = 'A2'

    # Tab 2: Bible Verses
    ws2 = wb.create_sheet('Bible Verses')
    headers2 = ['Bible Ref', 'Bible Verse Text']
    for col_idx, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    for row_idx, (bible_ref, bible_text) in enumerate(bible_verses_data, 2):
        ws2.cell(row=row_idx, column=1, value=bible_ref)
        ws2.cell(row=row_idx, column=2, value=bible_text)

    # Auto-fit columns
    for col_idx in range(1, 3):
        auto_fit_column(ws2, col_idx)

    # Freeze top row
    ws2.freeze_panes = 'A2'

    # Tab 3: Allusion Language
    ws3 = wb.create_sheet('Allusion Language')
    headers3 = ['Bible Ref', 'BoM Ref', 'Matched Phrases', 'Bible Text', 'BoM Text']
    for col_idx, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    for row_idx, (bible_ref, bom_ref, matched_phrases, bible_text, bom_text) in enumerate(allusions_with_matches_data, 2):
        ws3.cell(row=row_idx, column=1, value=bible_ref)
        ws3.cell(row=row_idx, column=2, value=bom_ref)
        ws3.cell(row=row_idx, column=3, value=matched_phrases)
        ws3.cell(row=row_idx, column=4, value=bible_text)
        ws3.cell(row=row_idx, column=5, value=bom_text)

    # Auto-fit columns
    for col_idx in range(1, 6):
        auto_fit_column(ws3, col_idx)

    # Freeze top row
    ws3.freeze_panes = 'A2'

    # Save
    wb.save(output_path)


def main():
    print("Loading data files...")
    hardy_data = load_hardy_data()
    phrase_index = load_phrase_index()
    scriptures = load_lds_scriptures()

    print(f"Loaded {len(hardy_data['entries'])} Hardy entries")
    print(f"Loaded {len(scriptures)} verses from lds-scriptures.txt")

    # Collect BoM verses
    print("\nProcessing BoM verses...")
    bom_verses_raw = collect_bom_verses(hardy_data)
    bom_verses_sorted = get_sorted_bom_verses(bom_verses_raw)

    bom_verses_data = []
    for bom_ref, bible_ref, entry_type in bom_verses_sorted:
        # Get BoM verse text
        parsed = parse_bom_ref(bom_ref)
        bom_texts = []
        for book_id, chapter, verse in parsed:
            text = get_bom_verse_text(scriptures, book_id, chapter, verse)
            if text:
                bom_texts.append(text)
        bom_text = ' | '.join(bom_texts) if bom_texts else ''

        # Get phrase matches
        matches = get_phrase_matches_for_bom(phrase_index, bom_ref)
        phrases_str = ' | '.join(matches) if matches else ''

        bom_verses_data.append((entry_type, bible_ref, bom_ref, bom_text, phrases_str))

    print(f"Collected {len(bom_verses_data)} BoM verses (quotations + allusions)")

    # Collect Bible verses
    print("\nProcessing Bible verses...")
    bible_refs = collect_unique_bible_refs(hardy_data)

    bible_verses_data = []
    for bible_ref in bible_refs:
        # Parse the reference (handle ranges and multiple verses)
        # For now, simple approach: convert dots to colons and look up
        # But need to handle ranges like "Isaiah 52.1-3"
        bible_ref_colon = bible_ref.replace('.', ':')

        # Check if it's a range or multiple verses
        bible_texts = []

        # Split by comma first
        for part in bible_ref.split(','):
            part = part.strip()
            if '-' in part:
                # Range like "Isaiah 52.1-3"
                last_space = part.rfind(' ')
                if last_space == -1:
                    continue
                book = part[:last_space]
                range_part = part[last_space+1:]
                if '.' not in range_part:
                    continue
                chapter_str, verse_range = range_part.split('.', 1)
                try:
                    chapter = int(chapter_str)
                    range_parts = verse_range.split('-')
                    start_verse = int(range_parts[0])
                    end_verse = int(range_parts[1])
                    for v in range(start_verse, end_verse + 1):
                        text = scriptures.get((book, chapter, v))
                        if text:
                            bible_texts.append(text)
                except (ValueError, IndexError):
                    continue
            else:
                # Single verse
                text = get_bible_verse_text(scriptures, part)
                if text:
                    bible_texts.append(text)

        bible_text = ' | '.join(bible_texts) if bible_texts else ''
        bible_verses_data.append((bible_ref, bible_text))

    print(f"Collected {len(bible_verses_data)} unique Bible references")

    # Collect allusions with matches
    print("\nProcessing allusions with phrase matches...")
    allusions_raw = collect_allusions_with_matches(hardy_data, phrase_index)

    allusions_with_matches_data = []
    for bible_ref, bom_ref, matched_phrases in allusions_raw:
        # Get Bible text - handle comma-separated and range references
        bible_texts = []
        for part in bible_ref.split(','):
            part = part.strip()
            if '-' in part:
                # Range like "Isaiah 52.1-3"
                last_space = part.rfind(' ')
                if last_space == -1:
                    continue
                book = part[:last_space]
                range_part = part[last_space+1:]
                if '.' not in range_part:
                    continue
                chapter_str, verse_range = range_part.split('.', 1)
                try:
                    chapter = int(chapter_str)
                    range_parts = verse_range.split('-')
                    start_verse = int(range_parts[0])
                    end_verse = int(range_parts[1])
                    for v in range(start_verse, end_verse + 1):
                        text = scriptures.get((book, chapter, v))
                        if text:
                            bible_texts.append(text)
                except (ValueError, IndexError):
                    continue
            else:
                # Single verse
                part_colon = part.replace('.', ':')
                last_space = part_colon.rfind(' ')
                if last_space != -1:
                    book = part_colon[:last_space]
                    verse_ref = part_colon[last_space+1:]
                    try:
                        chapter_str, verse_str = verse_ref.split(':', 1)
                        chapter = int(chapter_str)
                        verse = int(verse_str)
                        text = scriptures.get((book, chapter, verse))
                        if text:
                            bible_texts.append(text)
                    except (ValueError, AttributeError):
                        continue

        bible_text = ' | '.join(bible_texts) if bible_texts else ''

        # Get BoM text
        parsed = parse_bom_ref(bom_ref)
        bom_texts = []
        for book_id, chapter, verse in parsed:
            text = get_bom_verse_text(scriptures, book_id, chapter, verse)
            if text:
                bom_texts.append(text)
        bom_text = ' | '.join(bom_texts) if bom_texts else ''

        matched_phrases_str = ' | '.join(matched_phrases)

        allusions_with_matches_data.append((bible_ref, bom_ref, matched_phrases_str, bible_text, bom_text))

    print(f"Collected {len(allusions_with_matches_data)} allusions with phrase matches")

    # Create Excel file
    print("\nCreating Excel file...")
    output_path = 'data/allusion_analysis.xlsx'
    create_excel(output_path, bom_verses_data, bible_verses_data, allusions_with_matches_data)

    print(f"\nDone! Output written to {output_path}")
    print(f"\nStatistics:")
    print(f"  BoM Verses tab: {len(bom_verses_data)} rows")
    print(f"  Bible Verses tab: {len(bible_verses_data)} rows")
    print(f"  Allusion Language tab: {len(allusions_with_matches_data)} rows")


if __name__ == '__main__':
    main()
